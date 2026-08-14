from __future__ import print_function

import json
import re
import sys
import zipfile
from collections import OrderedDict
from pathlib import Path
from xml.etree import ElementTree


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE_PATH = PROJECT_ROOT / 'data' / 'publications.xlsx'
OUTPUT_PATH = PROJECT_ROOT / 'src' / 'data' / 'publications.js'
SOURCE_SHEET = 'Master_DB'

HEADERS = {
    'id': 'No.',
    'title': 'Title',
    'journal': 'Journal',
    'year': 'Year',
    'month': 'Month',
    'day': 'Day',
    'volume': 'Volume',
    'issue': 'Issue',
    'pages': 'Pages / Article No.',
    'authors': 'Authors',
    'doi': 'DOI',
    'doiUrl': 'DOI URL',
    'role': 'Role',
    'authorCount': 'Author Count',
    'authorOrder': 'Author Order',
    'researchAreas': 'Research Area (Draft)',
    'keywords': 'Keywords',
    'featured': 'Featured',
}

REQUIRED_FIELDS = ('id', 'title', 'journal', 'year')
TRUE_VALUES = {'yes', 'y', 'true', '1'}
FALSE_VALUES = {'', 'no', 'n', 'false', '0'}


class GenerationError(Exception):
    pass


def read_with_openpyxl(path):
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except (IOError, OSError, PermissionError) as error:
        raise GenerationError(
            'Cannot read data/publications.xlsx. Close Excel and try again. ({})'.format(error)
        )

    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise GenerationError('Excel sheet "{}" was not found.'.format(SOURCE_SHEET))
        worksheet = workbook[SOURCE_SHEET]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def column_index(reference):
    letters = re.match(r'^[A-Z]+', reference).group(0)
    result = 0
    for letter in letters:
        result = result * 26 + ord(letter) - ord('A') + 1
    return result - 1


def xml_text(node, namespace):
    return ''.join(part.text or '' for part in node.findall('.//{%s}t' % namespace))


def read_with_standard_library(path):
    main_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    package_rel_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'

    try:
        archive = zipfile.ZipFile(str(path), 'r')
    except (IOError, OSError, PermissionError, zipfile.BadZipFile) as error:
        raise GenerationError(
            'Cannot read data/publications.xlsx. Close Excel and try again. ({})'.format(error)
        )

    with archive:
        workbook = ElementTree.fromstring(archive.read('xl/workbook.xml'))
        relationships = ElementTree.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
        relationship_targets = {
            item.attrib['Id']: item.attrib['Target']
            for item in relationships.findall('{%s}Relationship' % package_rel_ns)
        }

        sheet_path = None
        for sheet in workbook.findall('.//{%s}sheet' % main_ns):
            if sheet.attrib.get('name') == SOURCE_SHEET:
                relationship_id = sheet.attrib.get('{%s}id' % rel_ns)
                target = relationship_targets.get(relationship_id)
                if target:
                    sheet_path = target.lstrip('/')
                    if not sheet_path.startswith('xl/'):
                        sheet_path = 'xl/' + sheet_path
                break

        if not sheet_path:
            raise GenerationError('Excel sheet "{}" was not found.'.format(SOURCE_SHEET))

        shared_strings = []
        if 'xl/sharedStrings.xml' in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read('xl/sharedStrings.xml'))
            shared_strings = [xml_text(item, main_ns) for item in shared_root.findall('{%s}si' % main_ns)]

        sheet = ElementTree.fromstring(archive.read(sheet_path))
        rows = []
        for row_node in sheet.findall('.//{%s}sheetData/{%s}row' % (main_ns, main_ns)):
            row_values = []
            for cell in row_node.findall('{%s}c' % main_ns):
                index = column_index(cell.attrib['r'])
                while len(row_values) <= index:
                    row_values.append(None)

                cell_type = cell.attrib.get('t')
                value_node = cell.find('{%s}v' % main_ns)
                if cell_type == 'inlineStr':
                    value = xml_text(cell, main_ns)
                elif value_node is None or value_node.text is None:
                    value = None
                elif cell_type == 's':
                    value = shared_strings[int(value_node.text)]
                elif cell_type == 'b':
                    value = value_node.text == '1'
                elif cell_type in ('str', 'e'):
                    value = value_node.text
                else:
                    number = float(value_node.text)
                    value = int(number) if number.is_integer() else number
                row_values[index] = value
            rows.append(row_values)
        return rows


def load_rows(path):
    rows = read_with_openpyxl(path)
    if rows is not None:
        print('Reading data/publications.xlsx with openpyxl.')
        return rows

    print('openpyxl is not installed; using the built-in XLSX reader.')
    return read_with_standard_library(path)


def empty_to_none(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def to_integer(value, field, excel_row):
    value = empty_to_none(value)
    if value is None:
        return None
    try:
        number = float(value)
        if not number.is_integer():
            raise ValueError
        return int(number)
    except (TypeError, ValueError):
        raise GenerationError('Excel row {} has an invalid {} value: {!r}'.format(excel_row, field, value))


def to_number(value, field, excel_row):
    value = empty_to_none(value)
    if value is None:
        return None
    try:
        number = float(value)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        raise GenerationError('Excel row {} has an invalid {} value: {!r}'.format(excel_row, field, value))


def to_text(value):
    value = empty_to_none(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def to_featured(value, excel_row):
    if isinstance(value, bool):
        return value
    normalized = '' if value is None else str(value).strip().lower()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise GenerationError('Excel row {} has an invalid Featured value: {!r}'.format(excel_row, value))


def convert_rows(rows):
    if not rows:
        raise GenerationError('The Master_DB sheet is empty.')

    header_positions = {}
    for index, value in enumerate(rows[0]):
        if value is not None:
            header_positions[str(value).strip()] = index

    missing_headers = [header for header in HEADERS.values() if header not in header_positions]
    if missing_headers:
        raise GenerationError('Missing required Excel column(s): {}'.format(', '.join(missing_headers)))

    publications = []
    doi_rows = {}
    for excel_row, row in enumerate(rows[1:], start=2):
        values = {
            field: row[header_positions[header]] if header_positions[header] < len(row) else None
            for field, header in HEADERS.items()
        }
        if all(empty_to_none(value) is None for value in values.values()):
            continue

        missing_fields = [field for field in REQUIRED_FIELDS if empty_to_none(values[field]) is None]
        if missing_fields:
            raise GenerationError(
                'Excel row {} is missing required field(s): {}'.format(excel_row, ', '.join(missing_fields))
            )

        doi = to_text(values['doi'])
        normalized_doi = doi.lower() if doi else None
        if normalized_doi in doi_rows:
            raise GenerationError(
                'Duplicate DOI {!r} found in Excel rows {} and {}.'.format(
                    doi, doi_rows[normalized_doi], excel_row
                )
            )
        if normalized_doi:
            doi_rows[normalized_doi] = excel_row

        authors_text = to_text(values['authors'])
        authors = [] if authors_text is None else [
            author.strip() for author in authors_text.split('/') if author.strip()
        ]
        research_area_text = to_text(values['researchAreas'])
        research_areas = [] if research_area_text is None else [
            area.strip() for area in research_area_text.split('/') if area.strip()
        ]
        keywords_text = to_text(values['keywords'])
        keywords = [] if keywords_text is None else [
            keyword.strip() for keyword in keywords_text.split('/') if keyword.strip()
        ]

        publications.append(OrderedDict([
            ('id', to_integer(values['id'], 'No.', excel_row)),
            ('year', to_integer(values['year'], 'Year', excel_row)),
            ('month', to_integer(values['month'], 'Month', excel_row)),
            ('day', to_integer(values['day'], 'Day', excel_row)),
            ('title', to_text(values['title'])),
            ('authors', authors),
            ('journal', to_text(values['journal'])),
            ('volume', to_text(values['volume'])),
            ('issue', to_text(values['issue'])),
            ('pages', to_text(values['pages'])),
            ('doi', doi),
            ('doiUrl', to_text(values['doiUrl'])),
            ('role', to_text(values['role'])),
            ('authorCount', to_integer(values['authorCount'], 'Author Count', excel_row)),
            ('authorOrder', to_integer(values['authorOrder'], 'Author Order', excel_row)),
            ('researchAreas', research_areas),
            ('keywords', keywords),
            ('featured', to_featured(values['featured'], excel_row)),
        ]))

    publications.sort(
        key=lambda publication: (
            publication['year'],
            publication['month'] if publication['month'] is not None else -1,
            publication['day'] if publication['day'] is not None else -1,
            publication['id'],
        ),
        reverse=True,
    )
    return publications


def write_javascript(publications, path):
    preamble = (
        '// AUTO-GENERATED FILE — DO NOT EDIT MANUALLY.\n'
        '// Source: data/publications.xlsx\n'
        '// Run `npm run publications` after editing the Excel database.\n\n'
    )
    content = '{}export const publications = {}\n\nexport default publications\n'.format(
        preamble,
        json.dumps(publications, ensure_ascii=False, indent=2),
    )
    temporary_path = path.with_suffix('.js.tmp')
    try:
        with temporary_path.open('w', encoding='utf-8', newline='\n') as output:
            output.write(content)
        temporary_path.replace(path)
    except (IOError, OSError, PermissionError) as error:
        raise GenerationError('Cannot write src/data/publications.js. ({})'.format(error))


def main():
    if not SOURCE_PATH.exists():
        raise GenerationError('Master database not found: data/publications.xlsx')
    publications = convert_rows(load_rows(SOURCE_PATH))
    write_javascript(publications, OUTPUT_PATH)
    print('Generated src/data/publications.js with {} publication records.'.format(len(publications)))


if __name__ == '__main__':
    try:
        main()
    except GenerationError as error:
        print('Publication generation failed: {}'.format(error), file=sys.stderr)
        sys.exit(1)
